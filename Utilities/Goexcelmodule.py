import openpyxl
from openpyxl.styles import PatternFill


def get_row_count(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return(sheet.max_row)

def get_column_count(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return (sheet.max_column)
def readdata(file,sheetname,row,column):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(row,column).value
def writwdata(file,sheetname,row,column,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row, column).value = data
    workbook.save(file)
def fillGreenColour(file,sheetname,row,column):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    greenfill=PatternFill(start_color='60b212',end_color='60b212',
                          fill_type='solid')
    sheet.cell(row,column).fill=greenfill
    workbook.save(file)
def fillRedColour(file,sheetname,row,column):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    redfill=PatternFill(start_color='ff0000',end_color='ff0000',
                          fill_type='solid')
    sheet.cell(row,column).fill=redfill
    workbook.save(file)



